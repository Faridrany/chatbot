"""
export_preprocessing.py
=======================
Script terpisah untuk mengekspor hasil preprocessing ke file Excel.

Cara pakai:
    python export_preprocessing.py
    python export_preprocessing.py --output hasil_preprocessing.xlsx

Sheet yang dihasilkan:
    0. Ringkasan
    1. Cleaning
    2. Casefolding
    3. Tokenizing
    4. Normalization
    5. Stopword Removal
    6. Stemming
    7. Final Result
"""

import os
import sys
import json
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================
# PATH — ikuti struktur folder project
# =====================================================
CURRENT_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(CURRENT_DIR, ".."))

DATA_DIR       = os.path.join(PROJECT_ROOT, "data")
PROCESSED_DIR  = os.path.join(DATA_DIR, "processed")
RAW_DIR        = os.path.join(DATA_DIR, "raw")
EXPORT_DIR     = os.path.join(DATA_DIR, "export")

CLEANED_PATH          = os.path.join(PROCESSED_DIR, "cleaned.json")
CASEFOLDED_PATH       = os.path.join(PROCESSED_DIR, "casefolded.json")
TOKENIZED_PATH        = os.path.join(PROCESSED_DIR, "tokenized.json")
NORMALIZED_PATH       = os.path.join(PROCESSED_DIR, "normalized.json")
STOP_REMOVED_PATH     = os.path.join(PROCESSED_DIR, "stop_removed.json")
STEMMED_PATH          = os.path.join(PROCESSED_DIR, "stemmed.json")
FINAL_PROCESSED_PATH  = os.path.join(PROCESSED_DIR, "final_processed.json")
DATASET_BERLABEL_PATH = os.path.join(RAW_DIR,        "dataset_berlabel.json")

DEFAULT_OUTPUT = os.path.join(EXPORT_DIR, "preprocessing_result.xlsx")

# =====================================================
# HELPER — load JSON
# =====================================================
def load_json(path):
    if not os.path.exists(path):
        print(f"  ⚠️  File tidak ditemukan, dilewati: {path}")
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

# =====================================================
# HELPER — style Excel
# =====================================================
COLOR_GREEN_DARK  = "2E7D32"
COLOR_GREEN_MID   = "4CAF50"
COLOR_GREEN_LIGHT = "A5D6A7"
COLOR_GREEN_PALE  = "F1F8E9"
COLOR_WHITE       = "FFFFFF"

def make_header_style():
    return {
        "font":      Font(name="Calibri", bold=True, color=COLOR_WHITE, size=11),
        "fill":      PatternFill("solid", fgColor=COLOR_GREEN_DARK),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":    _thin_border(),
    }

def make_alt_style():
    return {
        "fill":      PatternFill("solid", fgColor=COLOR_GREEN_PALE),
        "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border":    _thin_border(),
    }

def make_normal_style():
    return {
        "fill":      PatternFill("solid", fgColor=COLOR_WHITE),
        "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border":    _thin_border(),
    }

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_style(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)

def style_row(ws, row_num, col_count, is_header=False, alternate=False):
    if is_header:
        style = make_header_style()
    elif alternate:
        style = make_alt_style()
    else:
        style = make_normal_style()
    for col in range(1, col_count + 1):
        apply_style(ws.cell(row=row_num, column=col), style)

def set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_freeze_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def center(ws, row, col):
    ws.cell(row=row, column=col).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

# =====================================================
# SHEET BUILDERS
# =====================================================
def build_sheet_ringkasan(wb, steps_info: list):
    ws = wb.active
    ws.title = "Ringkasan"

    # Judul
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = "Laporan Hasil Preprocessing Teks"
    c.font      = Font(name="Calibri", bold=True, size=14, color=COLOR_WHITE)
    c.fill      = PatternFill("solid", fgColor=COLOR_GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Header tabel
    headers = ["No", "Tahap Preprocessing", "Deskripsi Proses", "Jumlah Data"]
    ws.append(headers)
    style_row(ws, 2, len(headers), is_header=True)
    ws.row_dimensions[2].height = 20

    for i, (tahap, desc, jml) in enumerate(steps_info, start=1):
        ws.append([i, tahap, desc, jml])
        style_row(ws, i + 2, len(headers), alternate=(i % 2 == 0))
        center(ws, i + 2, 1)
        center(ws, i + 2, 4)

    set_col_widths(ws, [6, 28, 60, 15])


def build_sheet_cleaning(wb, data: list):
    ws = wb.create_sheet("1. Cleaning")
    headers = ["No", "Teks Asli (Deskripsi)", "Hasil Cleaning"]
    ws.append(headers)
    style_row(ws, 1, len(headers), is_header=True)

    for i, row in enumerate(data, start=1):
        ws.append([i, row.get("deskripsi", ""), row.get("cleaned", row.get("hasil", ""))])
        style_row(ws, i + 1, len(headers), alternate=(i % 2 == 0))
        center(ws, i + 1, 1)

    set_col_widths(ws, [6, 65, 65])
    add_freeze_filter(ws)


def build_sheet_casefolding(wb, data: list):
    ws = wb.create_sheet("2. Casefolding")
    headers = ["No", "Teks Asli (Deskripsi)", "Hasil Casefolding"]
    ws.append(headers)
    style_row(ws, 1, len(headers), is_header=True)

    for i, row in enumerate(data, start=1):
        ws.append([i, row.get("deskripsi", ""), row.get("casefolded", row.get("hasil", ""))])
        style_row(ws, i + 1, len(headers), alternate=(i % 2 == 0))
        center(ws, i + 1, 1)

    set_col_widths(ws, [6, 65, 65])
    add_freeze_filter(ws)


def build_sheet_tokenizing(wb, data: list):
    ws = wb.create_sheet("3. Tokenizing")
    headers = ["No", "Teks Asli (Deskripsi)", "Hasil Token (List)", "Jumlah Token"]
    ws.append(headers)
    style_row(ws, 1, len(headers), is_header=True)

    for i, row in enumerate(data, start=1):
        tokens = row.get("tokenized", row.get("hasil", []))
        if isinstance(tokens, str):
            tokens = tokens.split()
        ws.append([i, row.get("deskripsi", ""), str(tokens), len(tokens)])
        style_row(ws, i + 1, len(headers), alternate=(i % 2 == 0))
        center(ws, i + 1, 1)
        center(ws, i + 1, 4)

    set_col_widths(ws, [6, 60, 65, 14])
    add_freeze_filter(ws)


def build_sheet_normalization(wb, tokenized: list, normalized: list):
    ws = wb.create_sheet("4. Normalization")
    headers = ["No", "Teks Asli (Deskripsi)", "Sebelum Normalisasi", "Sesudah Normalisasi", "Kata Diubah"]
    ws.append(headers)
    style_row(ws, 1, len(headers), is_header=True)

    for i, (tk, nm) in enumerate(zip(tokenized, normalized), start=1):
        before = tk.get("tokenized", tk.get("hasil", []))
        after  = nm.get("normalized", nm.get("hasil", []))
        if isinstance(before, str):
            before = before.split()
        if isinstance(after, str):
            after = after.split()

        # Cari kata yang berubah
        changed = []
        for b, a in zip(before, after):
            if b != a:
                changed.append(f"{b}→{a}")

        ws.append([
            i,
            nm.get("deskripsi", ""),
            " ".join(before),
            " ".join(after),
            ", ".join(changed) if changed else "-",
        ])
        style_row(ws, i + 1, len(headers), alternate=(i % 2 == 0))
        center(ws, i + 1, 1)

    set_col_widths(ws, [6, 55, 50, 50, 30])
    add_freeze_filter(ws)


def build_sheet_stopword(wb, normalized: list, stop_removed: list):
    ws = wb.create_sheet("5. Stopword Removal")
    headers = ["No", "Teks Asli (Deskripsi)", "Sebelum Stopword", "Sesudah Stopword", "Kata Dihapus", "Jumlah Dihapus"]
    ws.append(headers)
    style_row(ws, 1, len(headers), is_header=True)

    for i, (nm, sw) in enumerate(zip(normalized, stop_removed), start=1):
        before = nm.get("normalized", nm.get("hasil", []))
        after  = sw.get("stop_removed", sw.get("hasil", []))
        if isinstance(before, str):
            before = before.split()
        if isinstance(after, str):
            after = after.split()

        after_set = set(after)
        removed   = [w for w in before if w not in after_set]

        ws.append([
            i,
            sw.get("deskripsi", ""),
            " ".join(before),
            " ".join(after),
            ", ".join(removed) if removed else "-",
            len(removed),
        ])
        style_row(ws, i + 1, len(headers), alternate=(i % 2 == 0))
        center(ws, i + 1, 1)
        center(ws, i + 1, 6)

    set_col_widths(ws, [6, 50, 50, 50, 35, 16])
    add_freeze_filter(ws)


def build_sheet_stemming(wb, stop_removed: list, stemmed: list):
    ws = wb.create_sheet("6. Stemming")
    headers = ["No", "Teks Asli (Deskripsi)", "Sebelum Stemming", "Sesudah Stemming", "Jumlah Token Akhir"]
    ws.append(headers)
    style_row(ws, 1, len(headers), is_header=True)

    for i, (sw, st) in enumerate(zip(stop_removed, stemmed), start=1):
        before = sw.get("stop_removed", sw.get("hasil", []))
        after  = st.get("stemmed", st.get("hasil", []))
        if isinstance(before, str):
            before = before.split()
        if isinstance(after, str):
            after = after.split()

        ws.append([
            i,
            st.get("deskripsi", ""),
            " ".join(before),
            " ".join(after),
            len(after),
        ])
        style_row(ws, i + 1, len(headers), alternate=(i % 2 == 0))
        center(ws, i + 1, 1)
        center(ws, i + 1, 5)

    set_col_widths(ws, [6, 55, 55, 55, 18])
    add_freeze_filter(ws)


def build_sheet_final(wb, final_data: list):
    ws = wb.create_sheet("7. Final Result")
    headers = ["No", "Nama", "No WA", "Deskripsi Asli", "Kategori", "Teks Final (Input Model)"]
    ws.append(headers)
    style_row(ws, 1, len(headers), is_header=True)

    for i, item in enumerate(final_data, start=1):
        kategori = (
            item.get("kategori_prediksi")
            or item.get("Kategori")
            or item.get("kategori")
            or "-"
        )
        ws.append([
            i,
            item.get("nama", ""),
            str(item.get("no_wa", "")).replace("@c.us", ""),
            item.get("deskripsi", ""),
            kategori,
            item.get("final_text", ""),
        ])
        style_row(ws, i + 1, len(headers), alternate=(i % 2 == 0))
        center(ws, i + 1, 1)

    set_col_widths(ws, [6, 20, 18, 60, 18, 60])
    add_freeze_filter(ws)


# =====================================================
# MAIN EXPORT FUNCTION
# =====================================================
def export_to_excel(output_path: str = DEFAULT_OUTPUT):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print("📂 Membaca file hasil preprocessing...")
    cleaned_list      = load_json(CLEANED_PATH)
    casefolded_list   = load_json(CASEFOLDED_PATH)
    tokenized_list    = load_json(TOKENIZED_PATH)
    normalized_list   = load_json(NORMALIZED_PATH)
    stop_removed_list = load_json(STOP_REMOVED_PATH)
    stemmed_list      = load_json(STEMMED_PATH)
    final_list        = load_json(FINAL_PROCESSED_PATH)

    # Validasi minimal ada data
    if not final_list:
        print("❌ final_processed.json kosong atau tidak ditemukan.")
        print("   Jalankan: python main.py --train")
        sys.exit(1)

    print(f"   ✅ Cleaning      : {len(cleaned_list)} baris")
    print(f"   ✅ Casefolding   : {len(casefolded_list)} baris")
    print(f"   ✅ Tokenizing    : {len(tokenized_list)} baris")
    print(f"   ✅ Normalization : {len(normalized_list)} baris")
    print(f"   ✅ Stopword      : {len(stop_removed_list)} baris")
    print(f"   ✅ Stemming      : {len(stemmed_list)} baris")
    print(f"   ✅ Final Result  : {len(final_list)} baris")

    steps_info = [
        ("1. Cleaning",         "Hapus URL, karakter non-alfabet, dan spasi berlebih",       len(cleaned_list)),
        ("2. Casefolding",      "Ubah semua huruf menjadi huruf kecil (lowercase)",          len(casefolded_list)),
        ("3. Tokenizing",       "Pecah teks menjadi daftar token (kata per kata)",           len(tokenized_list)),
        ("4. Normalization",    "Ganti singkatan/slang dengan kata baku (gk→tidak, dll)",   len(normalized_list)),
        ("5. Stopword Removal", "Hapus kata umum yang tidak bermakna (dan, yang, di, dll)", len(stop_removed_list)),
        ("6. Stemming",         "Ubah kata ke bentuk dasar menggunakan Sastrawi",            len(stemmed_list)),
        ("7. Final Result",     "Teks bersih siap digunakan sebagai input model ML",         len(final_list)),
    ]

    print("\n📊 Membuat file Excel...")
    wb = Workbook()

    build_sheet_ringkasan(wb, steps_info)
    build_sheet_cleaning(wb, cleaned_list)
    build_sheet_casefolding(wb, casefolded_list)
    build_sheet_tokenizing(wb, tokenized_list)
    build_sheet_normalization(wb, tokenized_list, normalized_list)
    build_sheet_stopword(wb, normalized_list, stop_removed_list)
    build_sheet_stemming(wb, stop_removed_list, stemmed_list)
    build_sheet_final(wb, final_list)

    wb.save(output_path)
    print(f"\n✅ File Excel berhasil disimpan!")
    print(f"   📁 {output_path}")
    print(f"   📋 Sheet: Ringkasan | 1.Cleaning | 2.Casefolding | 3.Tokenizing | 4.Normalization | 5.Stopword | 6.Stemming | 7.Final Result")


# =====================================================
# ENTRY POINT
# =====================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Export hasil preprocessing teks ke file Excel"
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=DEFAULT_OUTPUT,
        help=f"Path file output .xlsx (default: {DEFAULT_OUTPUT})"
    )
    args = parser.parse_args()

    print("=== Export Preprocessing → Excel ===")
    print(f"Output : {args.output}")
    print()
    export_to_excel(args.output)
