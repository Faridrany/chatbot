# -*- coding: utf-8 -*-
import os
import re
import sys
import json
import joblib
import numpy as np
import pandas as pd
import hashlib
from datetime import datetime
from collections import Counter
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_selection import SelectPercentile, chi2
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score,
    f1_score, classification_report, confusion_matrix
)
from sklearn.preprocessing import LabelEncoder
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fix encoding Windows (cmd/PowerShell tidak support UTF-8 emoji)
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# =====================================================
# 1. PATH
# =====================================================
print("=== Sistem Klasifikasi Pengaduan Masyarakat ===")

CURRENT_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(CURRENT_DIR, ".."))

DATA_DIR       = os.path.join(PROJECT_ROOT, "data")
RAW_DIR        = os.path.join(DATA_DIR, "raw")
PROCESSED_DIR  = os.path.join(DATA_DIR, "processed")
FEATURE_DIR    = os.path.join(DATA_DIR, "features")
PREDICTION_DIR = os.path.join(DATA_DIR, "predictions")
MODEL_DIR      = os.path.join(PROJECT_ROOT, "model")
EXPORT_DIR     = os.path.join(DATA_DIR, "export")
STAGES_DIR     = os.path.join(DATA_DIR, "stages")  # Folder untuk file tahapan terpisah

for p in [RAW_DIR, PROCESSED_DIR, FEATURE_DIR, PREDICTION_DIR, MODEL_DIR, EXPORT_DIR, STAGES_DIR]:
    os.makedirs(p, exist_ok=True)

DATASET_BERLABEL_PATH = os.path.join(RAW_DIR,       "dataset_berlabel.json")
DATA_BARU_PATH        = os.path.join(RAW_DIR,       "data_baru.json")
CLEANED_PATH          = os.path.join(PROCESSED_DIR, "cleaned.json")
CASEFOLDED_PATH       = os.path.join(PROCESSED_DIR, "casefolded.json")
TOKENIZED_PATH        = os.path.join(PROCESSED_DIR, "tokenized.json")
NORMALIZED_PATH       = os.path.join(PROCESSED_DIR, "normalized.json")
STOP_REMOVED_PATH     = os.path.join(PROCESSED_DIR, "stop_removed.json")
STEMMED_PATH          = os.path.join(PROCESSED_DIR, "stemmed.json")
FINAL_PROCESSED_PATH  = os.path.join(PROCESSED_DIR, "final_processed.json")
MODEL_PATH            = os.path.join(MODEL_DIR,     "random_forest_model.pkl")
VECTORIZER_PATH       = os.path.join(MODEL_DIR,     "tfidf_vectorizer.pkl")
SELECTOR_PATH         = os.path.join(MODEL_DIR,     "feature_selector.pkl")
LABEL_ENCODER_PATH    = os.path.join(MODEL_DIR,     "label_encoder.pkl")
TFIDF_FEATURES_PATH   = os.path.join(FEATURE_DIR,  "tfidf_features.pkl")  # legacy, tidak digunakan
HASIL_PREDIKSI_PATH   = os.path.join(PREDICTION_DIR,"hasil_prediksi.json")
EXPORT_EXCEL_PATH     = os.path.join(EXPORT_DIR,    "preprocessing_result.xlsx")

# Stage-specific files (normalized structure with kode_pengaduan as key)
STAGE_TOKENISASI_PATH    = os.path.join(STAGES_DIR, "tokenisasi.json")
STAGE_TFIDF_PATH         = os.path.join(STAGES_DIR, "tfidf.json")
STAGE_FILTERING_PATH     = os.path.join(STAGES_DIR, "filtering.json")
STAGE_SELEKSI_FITUR_PATH = os.path.join(STAGES_DIR, "seleksi_fitur.json")
STAGE_RANDOM_FOREST_PATH = os.path.join(STAGES_DIR, "random_forest.json")
STAGE_OOB_PATH           = os.path.join(STAGES_DIR, "oob.json")
STAGE_CV_PATH            = os.path.join(STAGES_DIR, "cross_validation.json")
STAGE_GINI_PATH          = os.path.join(STAGES_DIR, "gini_splitting.json")
STAGE_GINI_SAMPLES_DIR   = os.path.join(STAGES_DIR, "gini_samples")
STAGE_VOTING_PATH        = os.path.join(STAGES_DIR, "majority_voting.json")

# =====================================================
# 2. SASTRAWI
# =====================================================
stemmer          = StemmerFactory().create_stemmer()
stopword_remover = StopWordRemoverFactory().create_stop_word_remover()

NORMALIZATION_DICT = {
    "gk": "tidak", "nggak": "tidak", "tdk": "tidak",
    "rt": "rukun tetangga", "rw": "rukun warga",
    "pju": "penerangan jalan umum"
}

# =====================================================
# 3. UTILITAS
# =====================================================
def save_json(data, filename):
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def load_json(filename):
    if not os.path.exists(filename):
        raise FileNotFoundError(f"File tidak ditemukan: {filename}")
    with open(filename, "r", encoding="utf-8") as f:
        return json.load(f)

def generate_kode_pengaduan(index, item):
    """
    Generate unique kode_pengaduan for each data entry.
    Format: PGD-XXXX (4-digit zero-padded sequential ID)
    """
    return f"PGD-{index + 1:04d}"

def get_pengaduan_detail(kode_pengaduan):
    """
    Helper function untuk mengambil semua detail satu pengaduan berdasarkan kode_pengaduan.
    Menggabungkan data dari processed.json + semua file tahap (stages).
    
    Args:
        kode_pengaduan: Kode unik pengaduan (e.g., "PGD-0001")
    
    Returns:
        dict: Complete details including all stages, or None if not found
    """
    try:
        # Load main processed data
        if not os.path.exists(FINAL_PROCESSED_PATH):
            return None
        
        processed_data = load_json(FINAL_PROCESSED_PATH)
        pengaduan_main = next((item for item in processed_data if item.get("kode_pengaduan") == kode_pengaduan), None)
        
        if not pengaduan_main:
            return None
        
        # Build complete detail object
        result = pengaduan_main.copy()
        
        # Add tokenization details
        if os.path.exists(STAGE_TOKENISASI_PATH):
            tokenisasi_data = load_json(STAGE_TOKENISASI_PATH)
            result["tokenisasi"] = tokenisasi_data.get(kode_pengaduan, {})
        
        # Add TF-IDF details
        if os.path.exists(STAGE_TFIDF_PATH):
            tfidf_data = load_json(STAGE_TFIDF_PATH)
            result["tfidf"] = tfidf_data.get(kode_pengaduan, {})
        
        # Add filtering details
        if os.path.exists(STAGE_FILTERING_PATH):
            filtering_data = load_json(STAGE_FILTERING_PATH)
            result["filtering"] = filtering_data.get(kode_pengaduan, {})
        
        # Add seleksi fitur details
        if os.path.exists(STAGE_SELEKSI_FITUR_PATH):
            seleksi_data = load_json(STAGE_SELEKSI_FITUR_PATH)
            result["seleksi_fitur"] = seleksi_data.get(kode_pengaduan, {})
        
        # Add random forest details
        if os.path.exists(STAGE_RANDOM_FOREST_PATH):
            rf_data = load_json(STAGE_RANDOM_FOREST_PATH)
            result["random_forest"] = rf_data.get(kode_pengaduan, {})
        
        return result
        
    except Exception as e:
        print(f"[ERROR] get_pengaduan_detail({kode_pengaduan}): {e}")
        return None

# =====================================================
# 4. PREPROCESSING
# Urutan: Cleaning → Casefolding → Tokenizing →
#         Normalization → Stopword Removal → Stemming
# =====================================================
def cleaning(text):
    """Hapus URL, karakter non-alfabet, dan spasi berlebih."""
    text = re.sub(r"http\S+|www\S+", "", text)
    text = re.sub(r"[^a-zA-Z\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def case_folding(text):
    """Ubah semua huruf menjadi huruf kecil."""
    return text.lower()

def tokenization(text):
    """Pecah teks menjadi daftar token (kata)."""
    return text.split()

def normalization(tokens):
    """Ganti singkatan/slang dengan kata baku."""
    return [NORMALIZATION_DICT.get(t, t) for t in tokens]

def stopword_removal(tokens):
    """Hapus kata-kata umum yang tidak bermakna (stopword)."""
    return stopword_remover.remove(" ".join(tokens)).split()

def stemming(tokens):
    """Ubah kata ke bentuk dasar (stem)."""
    return stemmer.stem(" ".join(tokens)).split()

def preprocess_pipeline(data, save_files=True, append_files=False, start_index=0):
    """
    Pipeline preprocessing dengan urutan:
    1. Cleaning
    2. Casefolding
    3. Tokenizing
    4. Normalization
    5. Stopword Removal
    6. Stemming
    
    Args:
        data: List of data entries
        save_files: Whether to save intermediate files
        append_files: Whether to append to existing files
        start_index: Starting index for kode_pengaduan generation (for new data)
    
    Returns:
        final_data: List of processed data with kode_pengaduan
    """
    print("[*] Memulai proses preprocessing...")
    print("    Urutan: Cleaning -> Casefolding -> Tokenizing -> Normalization -> Stopword -> Stemming")

    cleaned_data      = []
    casefolded_data   = []
    tokenized_data    = []
    normalized_data   = []
    stop_removed_data = []
    stemmed_data      = []
    final_data        = []
    
    # Stage-specific data structures (will be saved to separate files)
    tokenisasi_dict = {}
    
    for i, item in enumerate(data):
        text = str(item.get("deskripsi", ""))
        
        # Generate kode_pengaduan
        kode = generate_kode_pengaduan(start_index + i, item)

        # Step 1: Cleaning
        cleaned = cleaning(text)
        cleaned_data.append({"kode_pengaduan": kode, "deskripsi": text, "cleaned": cleaned})

        # Step 2: Casefolding
        casefolded = case_folding(cleaned)
        casefolded_data.append({"kode_pengaduan": kode, "deskripsi": text, "casefolded": casefolded})

        # Step 3: Tokenizing
        tokens = tokenization(casefolded)
        tokenized_data.append({"kode_pengaduan": kode, "deskripsi": text, "tokenized": tokens})
        
        # Count term frequency for tokenization stage
        unigrams = [t for t in tokens if " " not in t]
        bigrams = []
        for j in range(len(tokens) - 1):
            bigrams.append(f"{tokens[j]} {tokens[j+1]}")
        
        unigram_freq = dict(Counter(unigrams))
        bigram_freq = dict(Counter(bigrams))
        
        tokenisasi_dict[kode] = {
            "unigram": unigram_freq,
            "bigram": bigram_freq,
            "total_tokens": len(tokens)
        }

        # Step 4: Normalization
        normalized = normalization(tokens)
        normalized_data.append({"kode_pengaduan": kode, "deskripsi": text, "normalized": normalized})

        # Step 5: Stopword Removal
        stop_removed = stopword_removal(normalized)
        stop_removed_data.append({"kode_pengaduan": kode, "deskripsi": text, "stop_removed": stop_removed})

        # Step 6: Stemming
        stemmed = stemming(stop_removed)
        stemmed_data.append({"kode_pengaduan": kode, "deskripsi": text, "stemmed": stemmed})

        final_text = " ".join(stemmed)
        final_item = item.copy()
        final_item["kode_pengaduan"] = kode
        final_item["final_text"] = final_text
        final_data.append(final_item)

    if save_files:
        def load_or_empty(path):
            if append_files and os.path.exists(path):
                return load_json(path)
            return []

        save_json(load_or_empty(CLEANED_PATH)         + cleaned_data,      CLEANED_PATH)
        save_json(load_or_empty(CASEFOLDED_PATH)      + casefolded_data,   CASEFOLDED_PATH)
        save_json(load_or_empty(TOKENIZED_PATH)       + tokenized_data,    TOKENIZED_PATH)
        save_json(load_or_empty(NORMALIZED_PATH)      + normalized_data,   NORMALIZED_PATH)
        save_json(load_or_empty(STOP_REMOVED_PATH)    + stop_removed_data, STOP_REMOVED_PATH)
        save_json(load_or_empty(STEMMED_PATH)         + stemmed_data,      STEMMED_PATH)
        
        # Save tokenization stage file (keyed by kode_pengaduan)
        if append_files and os.path.exists(STAGE_TOKENISASI_PATH):
            existing = load_json(STAGE_TOKENISASI_PATH)
            existing.update(tokenisasi_dict)
            save_json(existing, STAGE_TOKENISASI_PATH)
        else:
            save_json(tokenisasi_dict, STAGE_TOKENISASI_PATH)

    print("[OK] Preprocessing selesai.")
    return final_data

# =====================================================
# 5. EXPORT EXCEL
# =====================================================
def export_preprocessing_to_excel(data, final_data, output_path=None):
    """
    Export hasil setiap tahap preprocessing ke file Excel.
    Setiap tahap memiliki sheet tersendiri, plus sheet ringkasan.

    Args:
        data       : list data asli (dari dataset_berlabel.json)
        final_data : list hasil preprocessing (dari preprocess_pipeline)
        output_path: path file .xlsx (default: EXPORT_EXCEL_PATH)
    """
    if output_path is None:
        output_path = EXPORT_EXCEL_PATH

    print(f"[*] Mengekspor hasil preprocessing ke Excel: {output_path}")

    # Load tiap file JSON hasil preprocessing
    def load_step(path):
        if os.path.exists(path):
            return load_json(path)
        return []

    cleaned_list      = load_step(CLEANED_PATH)
    casefolded_list   = load_step(CASEFOLDED_PATH)
    tokenized_list    = load_step(TOKENIZED_PATH)
    normalized_list   = load_step(NORMALIZED_PATH)
    stop_removed_list = load_step(STOP_REMOVED_PATH)
    stemmed_list      = load_step(STEMMED_PATH)

    wb = Workbook()

    # ── Style helpers ──────────────────────────────────────────
    header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="2E7D32")   # hijau tua
    subhead_fill   = PatternFill("solid", fgColor="A5D6A7")   # hijau muda
    alt_fill       = PatternFill("solid", fgColor="F1F8E9")   # hijau sangat muda
    center_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border

    def style_data_row(ws, row_num, col_count, alternate=False):
        fill = alt_fill if alternate else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill      = fill
            cell.alignment = left_align
            cell.border    = thin_border

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def freeze_and_filter(ws, freeze_cell="A2"):
        ws.freeze_panes = freeze_cell
        ws.auto_filter.ref = ws.dimensions

    # ── Sheet 1: Ringkasan ─────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Ringkasan"

    ws_sum.merge_cells("A1:C1")
    title_cell = ws_sum["A1"]
    title_cell.value     = "Ringkasan Hasil Preprocessing"
    title_cell.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill      = header_fill
    title_cell.alignment = center_align

    ws_sum.append(["Tahap", "Deskripsi", "Jumlah Data"])
    style_header_row(ws_sum, 2, 3)

    steps_info = [
        ("1. Cleaning",          "Hapus URL, karakter non-alfabet, spasi berlebih",          len(cleaned_list)),
        ("2. Casefolding",       "Ubah semua huruf menjadi huruf kecil",                     len(casefolded_list)),
        ("3. Tokenizing",        "Pecah teks menjadi daftar token (kata)",                   len(tokenized_list)),
        ("4. Normalization",     "Ganti singkatan/slang dengan kata baku",                   len(normalized_list)),
        ("5. Stopword Removal",  "Hapus kata umum yang tidak bermakna",                      len(stop_removed_list)),
        ("6. Stemming",          "Ubah kata ke bentuk dasar",                                len(stemmed_list)),
        ("Final Result",         "Teks siap pakai untuk training/prediksi model",            len(final_data)),
    ]

    for i, (tahap, desc, jml) in enumerate(steps_info, start=3):
        ws_sum.append([tahap, desc, jml])
        style_data_row(ws_sum, i, 3, alternate=(i % 2 == 0))
        ws_sum.cell(row=i, column=3).alignment = Alignment(horizontal="center", vertical="center")

    set_col_widths(ws_sum, [25, 55, 15])
    ws_sum.row_dimensions[1].height = 30
    ws_sum.row_dimensions[2].height = 20

    # ── Sheet 2: Cleaning ──────────────────────────────────────
    ws_cl = wb.create_sheet("1. Cleaning")
    ws_cl.append(["No", "Teks Asli (Deskripsi)", "Hasil Cleaning"])
    style_header_row(ws_cl, 1, 3)
    for i, row in enumerate(cleaned_list, start=1):
        ws_cl.append([i, row.get("deskripsi", ""), row.get("cleaned", "")])
        style_data_row(ws_cl, i + 1, 3, alternate=(i % 2 == 0))
        ws_cl.cell(row=i + 1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    set_col_widths(ws_cl, [6, 60, 60])
    freeze_and_filter(ws_cl)

    # ── Sheet 3: Casefolding ───────────────────────────────────
    ws_cf = wb.create_sheet("2. Casefolding")
    ws_cf.append(["No", "Teks Asli (Deskripsi)", "Hasil Casefolding"])
    style_header_row(ws_cf, 1, 3)
    for i, row in enumerate(casefolded_list, start=1):
        ws_cf.append([i, row.get("deskripsi", ""), row.get("casefolded", "")])
        style_data_row(ws_cf, i + 1, 3, alternate=(i % 2 == 0))
        ws_cf.cell(row=i + 1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    set_col_widths(ws_cf, [6, 60, 60])
    freeze_and_filter(ws_cf)

    # ── Sheet 4: Tokenizing ────────────────────────────────────
    ws_tk = wb.create_sheet("3. Tokenizing")
    ws_tk.append(["No", "Teks Asli (Deskripsi)", "Hasil Tokenizing", "Jumlah Token"])
    style_header_row(ws_tk, 1, 4)
    for i, row in enumerate(tokenized_list, start=1):
        tokens = row.get("tokenized", [])
        ws_tk.append([i, row.get("deskripsi", ""), str(tokens), len(tokens)])
        style_data_row(ws_tk, i + 1, 4, alternate=(i % 2 == 0))
        ws_tk.cell(row=i + 1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws_tk.cell(row=i + 1, column=4).alignment = Alignment(horizontal="center", vertical="center")
    set_col_widths(ws_tk, [6, 60, 60, 14])
    freeze_and_filter(ws_tk)

    # ── Sheet 5: Normalization ─────────────────────────────────
    ws_nm = wb.create_sheet("4. Normalization")
    ws_nm.append(["No", "Teks Asli (Deskripsi)", "Sebelum Normalisasi", "Sesudah Normalisasi"])
    style_header_row(ws_nm, 1, 4)
    for i, (tk_row, nm_row) in enumerate(zip(tokenized_list, normalized_list), start=1):
        before = str(tk_row.get("tokenized", []))
        after  = str(nm_row.get("normalized", []))
        ws_nm.append([i, nm_row.get("deskripsi", ""), before, after])
        style_data_row(ws_nm, i + 1, 4, alternate=(i % 2 == 0))
        ws_nm.cell(row=i + 1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    set_col_widths(ws_nm, [6, 60, 55, 55])
    freeze_and_filter(ws_nm)

    # ── Sheet 6: Stopword Removal ──────────────────────────────
    ws_sw = wb.create_sheet("5. Stopword Removal")
    ws_sw.append(["No", "Teks Asli (Deskripsi)", "Sebelum Stopword", "Sesudah Stopword", "Kata Dihapus"])
    style_header_row(ws_sw, 1, 5)
    for i, (nm_row, sw_row) in enumerate(zip(normalized_list, stop_removed_list), start=1):
        before_tokens = set(nm_row.get("normalized", []))
        after_tokens  = set(sw_row.get("stop_removed", []))
        removed       = sorted(before_tokens - after_tokens)
        ws_sw.append([
            i,
            sw_row.get("deskripsi", ""),
            " ".join(nm_row.get("normalized", [])),
            " ".join(sw_row.get("stop_removed", [])),
            ", ".join(removed) if removed else "-",
        ])
        style_data_row(ws_sw, i + 1, 5, alternate=(i % 2 == 0))
        ws_sw.cell(row=i + 1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    set_col_widths(ws_sw, [6, 55, 50, 50, 35])
    freeze_and_filter(ws_sw)

    # ── Sheet 7: Stemming ──────────────────────────────────────
    ws_st = wb.create_sheet("6. Stemming")
    ws_st.append(["No", "Teks Asli (Deskripsi)", "Sebelum Stemming", "Sesudah Stemming"])
    style_header_row(ws_st, 1, 4)
    for i, (sw_row, st_row) in enumerate(zip(stop_removed_list, stemmed_list), start=1):
        ws_st.append([
            i,
            st_row.get("deskripsi", ""),
            " ".join(sw_row.get("stop_removed", [])),
            " ".join(st_row.get("stemmed", [])),
        ])
        style_data_row(ws_st, i + 1, 4, alternate=(i % 2 == 0))
        ws_st.cell(row=i + 1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    set_col_widths(ws_st, [6, 60, 55, 55])
    freeze_and_filter(ws_st)

    # ── Sheet 8: Final Result ──────────────────────────────────
    ws_fn = wb.create_sheet("Final Result")
    ws_fn.append(["No", "Nama", "No WA", "Deskripsi Asli", "Kategori", "Teks Final (Siap Model)"])
    style_header_row(ws_fn, 1, 6)
    for i, item in enumerate(final_data, start=1):
        kategori = item.get("kategori_prediksi") or item.get("Kategori") or item.get("kategori") or "-"
        ws_fn.append([
            i,
            item.get("nama", ""),
            str(item.get("no_wa", "")).replace("@c.us", ""),
            item.get("deskripsi", ""),
            kategori,
            item.get("final_text", ""),
        ])
        style_data_row(ws_fn, i + 1, 6, alternate=(i % 2 == 0))
        ws_fn.cell(row=i + 1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    set_col_widths(ws_fn, [6, 20, 18, 60, 18, 55])
    freeze_and_filter(ws_fn)

    wb.save(output_path)
    print(f"[OK] File Excel berhasil disimpan: {output_path}")
    print(f"     Sheet: Ringkasan | 1.Cleaning | 2.Casefolding | 3.Tokenizing | 4.Normalization | 5.Stopword | 6.Stemming | Final Result")
    return output_path

# =====================================================
# 6. TRAINING
# =====================================================
def train_model():
    print("[*] Memulai proses pelatihan model...")

    data = load_json(DATASET_BERLABEL_PATH)
    print(f"[*] Jumlah data latih: {len(data)}")

    data = preprocess_pipeline(data, save_files=True, append_files=False)

    df = pd.DataFrame(data)

    # DEBUG
    print(f"Total sebelum filter kosong : {len(df)}")
    df = df[df["final_text"].str.strip() != ""]
    print(f"Total sesudah filter kosong : {len(df)}")

    X_text    = df["final_text"]
    label_col = "Kategori" if "Kategori" in df.columns else "kategori"
    y         = df[label_col]

    print(f"Jumlah label unik           : {y.nunique()} → {sorted(y.unique())}")

    label_encoder = LabelEncoder()
    y_encoded     = label_encoder.fit_transform(y)

    # TF-IDF hanya unigram (1-gram)
    vectorizer = TfidfVectorizer(
        max_features=5000,
        min_df=2,
        max_df=0.95,
        ngram_range=(1, 1),
    )
    X_tfidf = vectorizer.fit_transform(X_text)
    print(f"Shape TF-IDF                : {X_tfidf.shape}")
    # TIDAK disimpan ke disk — matriks TF-IDF tidak diperlukan saat inferensi

    # Seleksi fitur dengan SelectPercentile + chi2 (ambil 80% fitur terbaik)
    selector   = SelectPercentile(score_func=chi2, percentile=80)
    X_selected = selector.fit_transform(X_tfidf, y_encoded)
    print(f"Shape setelah seleksi fitur : {X_selected.shape}")

    print(f"\nJumlah data awal  : {len(df)}")
    print(f"Jumlah final_text : {len(X_text)}")
    print(f"Jumlah label      : {len(y)}")

    # shuffle=True  → pastikan data diacak sebelum dipotong (hindari urutan kategori berurutan)
    # stratify=y_encoded → jamin proporsi tiap kelas sama rata di train & test (misal 80/20 per kelas)
    X_train, X_test, y_train, y_test = train_test_split(
        X_selected, y_encoded,
        test_size=0.2,
        random_state=42,
        stratify=y_encoded,
        shuffle=True,
    )
    print(f"Train             : {len(y_train)}")
    print(f"Test              : {len(y_test)}")

    # class_weight="balanced_subsample" → setiap pohon menghitung ulang bobot kelas
    # dari data bootstrap-nya masing-masing, bukan dari keseluruhan dataset.
    # Ini mencegah bias saat satu kelas kurang terwakili di kantong bootstrap tertentu.
    model = RandomForestClassifier(
        n_estimators=5,
        max_depth=None,
        min_samples_split=5,
        min_samples_leaf=2,
        class_weight="balanced_subsample",
        n_jobs=-1,
        random_state=42,
        oob_score=True,
    )
    model.fit(X_train, y_train)

    y_pred = model.predict(X_test)
    print(f"\nJumlah prediksi   : {len(y_pred)}")
    print(f"Jumlah label test : {len(y_test)}")

    print("\n=== HASIL EVALUASI MODEL ===")
    acc  = accuracy_score(y_test, y_pred)
    prec = precision_score(y_test, y_pred, average="weighted")
    rec  = recall_score(y_test, y_pred, average="weighted")
    f1   = f1_score(y_test, y_pred, average="weighted")
    print(f"Accuracy  : {acc}")
    print(f"Precision : {prec}")
    print(f"Recall    : {rec}")
    print(f"F1-Score  : {f1}")
    print(f"OOB Score : {model.oob_score_}")
    print("\nClassification Report:\n", classification_report(y_test, y_pred))
    print("\nConfusion Matrix:\n", confusion_matrix(y_test, y_pred))

    # Cross Validation 5-fold — simpan skor ringkasan untuk hasil_training.json
    print("\n[*] Menjalankan Cross Validation 5-fold (n_jobs=-1)...")
    cv_scores = cross_val_score(
        model, X_selected, y_encoded, cv=5, scoring="accuracy", n_jobs=-1
    )
    print(f"CV Scores : {cv_scores}")
    print(f"CV Mean   : {cv_scores.mean():.4f}  ±  {cv_scores.std():.4f}")

    joblib.dump(model,         MODEL_PATH)
    joblib.dump(vectorizer,    VECTORIZER_PATH)
    joblib.dump(selector,      SELECTOR_PATH)
    joblib.dump(label_encoder, LABEL_ENCODER_PATH)
    print(f"\n[OK] Model berhasil disimpan di: {MODEL_PATH}")

    # Simpan hasil evaluasi untuk frontend (halaman Statistik)
    # Hitung per-class metrics dari y_test (data test 120)
    cm = confusion_matrix(y_test, y_pred)
    classes = label_encoder.classes_.tolist()
    per_class = {}
    for idx, cat in enumerate(classes):
        tp = int(cm[idx, idx])
        fp = int(cm[:, idx].sum() - tp)
        fn = int(cm[idx, :].sum() - tp)
        prec_c = tp / (tp + fp) if (tp + fp) > 0 else 0
        rec_c  = tp / (tp + fn) if (tp + fn) > 0 else 0
        f1_c   = 2 * prec_c * rec_c / (prec_c + rec_c) if (prec_c + rec_c) > 0 else 0
        per_class[cat] = {
            "precision": round(prec_c, 4),
            "recall"   : round(rec_c,  4),
            "f1"       : round(f1_c,   4),
            "support"  : int(cm[idx, :].sum()),
            "tp": tp, "fp": fp, "fn": fn,
        }

    # Confusion matrix sebagai dict
    cm_dict = {}
    for i, actual in enumerate(classes):
        cm_dict[actual] = {}
        for j, pred_cat in enumerate(classes):
            cm_dict[actual][pred_cat] = int(cm[i, j])

    hasil_evaluasi = {
        "akurasi"      : round(acc,  4),
        "presisi"      : round(prec, 4),
        "recall"       : round(rec,  4),
        "f1_score"     : round(f1,   4),
        "oob_score"    : round(float(model.oob_score_), 4),
        "cv_mean"      : round(float(cv_scores.mean()), 4),
        "cv_std"       : round(float(cv_scores.std()),  4),
        "total_data"   : len(df),
        "data_train"   : len(y_train),
        "data_test"    : len(y_test),
        "jumlah_kelas" : int(y.nunique()),
        "kelas"        : sorted(y.unique().tolist()),
        "fitur_tfidf"  : int(X_tfidf.shape[1]),
        "fitur_selected": int(X_selected.shape[1]),
        "estimators"   : 5,
        "ngram_range"  : [1, 1],        "perClass"     : per_class,
        "confusionMatrix": cm_dict,
    }
    save_json(hasil_evaluasi, os.path.join(DATA_DIR, "hasil_training.json"))
    print("[OK] Hasil evaluasi disimpan ke: hasil_training.json")

    # ═══════════════════════════════════════════════════════════════
    # GENERATE OOB.JSON — data nganggur per pohon (20 pohon pertama)
    # ═══════════════════════════════════════════════════════════════
    print("\n[*] Generating oob.json (data aktual per pohon)...")

    # Rekonstruksi mapping train_idx → kode & label (sama dengan bootstrap.json)
    _df_reset_oob = df.reset_index(drop=True)
    _all_idx      = list(range(len(_df_reset_oob)))
    from sklearn.model_selection import train_test_split as _tts2
    _train_idx_oob, _ = _tts2(
        _all_idx, test_size=0.2, random_state=42, stratify=y_encoded, shuffle=True
    )
    _train_kode  = [_df_reset_oob.iloc[i]["kode_pengaduan"] for i in _train_idx_oob]
    _train_label = [label_encoder.inverse_transform([y_encoded[i]])[0] for i in _train_idx_oob]
    _n_train     = len(_train_idx_oob)

    oob_per_pohon   = {}
    oob_akumulasi   = {}  # kode → {muncul_di_pohon, prediksi_tiap_pohon}

    for tree_num in range(min(20, len(model.estimators_))):
        tree         = model.estimators_[tree_num]
        samp_idx     = model.estimators_samples_[tree_num]
        in_bag_set   = set(samp_idx.tolist())

        # Kumpulkan indeks OOB untuk pohon ini
        oob_indices = [i for i in range(_n_train) if i not in in_bag_set]

        if not oob_indices:
            continue

        # Buat sub-matriks X_train untuk prediksi OOB
        X_train_sub = X_train[oob_indices]
        oob_preds   = tree.predict(X_train_sub).astype(int)
        oob_labels  = label_encoder.inverse_transform(oob_preds)

        oob_data_list = []
        benar_count   = 0
        for rank, orig_idx in enumerate(oob_indices):
            if orig_idx >= _n_train:
                continue
            kode        = _train_kode[orig_idx]
            label_asli  = _train_label[orig_idx]
            pred_oob    = oob_labels[rank]
            is_benar    = (pred_oob == label_asli)
            if is_benar:
                benar_count += 1

            oob_data_list.append({
                "kode_pengaduan": kode,
                "label_asli"    : label_asli,
                "prediksi_oob"  : pred_oob,
                "benar"         : bool(is_benar),
            })

            # Akumulasi lintas pohon
            if kode not in oob_akumulasi:
                oob_akumulasi[kode] = {
                    "muncul_di_pohon"    : [],
                    "prediksi_tiap_pohon": [],
                }
            oob_akumulasi[kode]["muncul_di_pohon"].append(tree_num + 1)
            oob_akumulasi[kode]["prediksi_tiap_pohon"].append(pred_oob)

        tree_acc = round(benar_count / len(oob_data_list), 4) if oob_data_list else 0.0
        oob_per_pohon[f"tree_{tree_num + 1}"] = {
            "oob_count"   : len(oob_data_list),
            "benar"       : benar_count,
            "salah"       : len(oob_data_list) - benar_count,
            "oob_accuracy": tree_acc,
            "oob_data"    : oob_data_list,
        }

    # Finalisasi akumulasi — hitung prediksi final & konsistensi
    from collections import Counter as _C2
    akumulasi_final = {}
    for kode, v in oob_akumulasi.items():
        preds        = v["prediksi_tiap_pohon"]
        counter      = _C2(preds)
        pred_final   = counter.most_common(1)[0][0]
        konsisten    = len(counter) == 1
        # label asli bisa dicari dari _train_label via kode
        try:
            idx_asli = _train_kode.index(kode)
            label_a  = _train_label[idx_asli]
        except ValueError:
            label_a  = "-"
        akumulasi_final[kode] = {
            "muncul_di_pohon"    : v["muncul_di_pohon"],
            "prediksi_tiap_pohon": preds,
            "prediksi_final_oob" : pred_final,
            "label_asli"         : label_a,
            "konsisten"          : konsisten,
        }

    oob_output = {
        "oob_score_global": round(float(model.oob_score_), 4),
        "jumlah_pohon"    : min(20, len(model.estimators_)),
        "per_pohon"       : oob_per_pohon,
        "akumulasi"       : akumulasi_final,
    }
    save_json(oob_output, STAGE_OOB_PATH)
    print(f"  [OK] oob.json saved ({len(oob_per_pohon)} pohon, {len(akumulasi_final)} entri akumulasi)")

    # ═══════════════════════════════════════════════════════════════
    # GENERATE CROSS_VALIDATION.JSON — detail per fold (5-fold)
    # ═══════════════════════════════════════════════════════════════
    print("\n[*] Generating cross_validation.json (detail per fold)...")

    from sklearn.model_selection import StratifiedKFold
    skf     = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
    # Gunakan X_selected & y_encoded (data lengkap sebelum train-test split utama)
    all_kode_list = _df_reset_oob["kode_pengaduan"].tolist()

    cv_folds       = {}
    error_tracker  = {}  # kode → {fold_idx yang salah, prediksi di fold itu}

    cv_scores_list = []
    for fold_idx, (cv_train_idx, cv_test_idx) in enumerate(skf.split(X_selected, y_encoded)):
        fold_num = fold_idx + 1

        X_cv_train = X_selected[cv_train_idx]
        X_cv_test  = X_selected[cv_test_idx]
        y_cv_train = y_encoded[cv_train_idx]
        y_cv_test  = y_encoded[cv_test_idx]

        # Train model khusus fold ini
        cv_model = RandomForestClassifier(
            n_estimators=5,
            max_depth=30,
            min_samples_split=5,
            min_samples_leaf=2,
            class_weight="balanced_subsample",
            n_jobs=-1,
            random_state=42 + fold_idx,
        )
        cv_model.fit(X_cv_train, y_cv_train)
        y_cv_pred = cv_model.predict(X_cv_test)

        fold_acc    = round(float(accuracy_score(y_cv_test, y_cv_pred)), 4)
        cv_scores_list.append(fold_acc)

        training_ids = [all_kode_list[i] for i in cv_train_idx if i < len(all_kode_list)]
        testing_ids  = [all_kode_list[i] for i in cv_test_idx  if i < len(all_kode_list)]

        hasil_testing = []
        for rank, orig_idx in enumerate(cv_test_idx):
            if orig_idx >= len(all_kode_list):
                continue
            kode       = all_kode_list[orig_idx]
            label_asli = label_encoder.inverse_transform([y_cv_test[rank]])[0]
            prediksi   = label_encoder.inverse_transform([y_cv_pred[rank]])[0]
            is_benar   = (prediksi == label_asli)
            hasil_testing.append({
                "kode_pengaduan": kode,
                "label_asli"    : label_asli,
                "prediksi"      : prediksi,
                "benar"         : bool(is_benar),
            })
            if not is_benar:
                if kode not in error_tracker:
                    error_tracker[kode] = {
                        "label_asli"  : label_asli,
                        "salah_di_fold": [],
                        "prediksi_di_fold": [],
                    }
                error_tracker[kode]["salah_di_fold"].append(fold_num)
                error_tracker[kode]["prediksi_di_fold"].append(prediksi)

        cv_folds[f"fold_{fold_num}"] = {
            "training_size": len(training_ids),
            "testing_size" : len(testing_ids),
            "training_ids" : training_ids,
            "testing_ids"  : testing_ids,
            "hasil_testing": hasil_testing,
            "akurasi"      : fold_acc,
        }
        print(f"    Fold {fold_num}: acc={fold_acc:.4f}")

    # Data yang konsisten salah di >= 3 fold
    data_konsisten_salah = []
    for kode, v in error_tracker.items():
        if len(v["salah_di_fold"]) >= 3:
            from collections import Counter as _C3
            pred_counter   = _C3(v["prediksi_di_fold"])
            pred_dominan   = pred_counter.most_common(1)[0][0]
            data_konsisten_salah.append({
                "kode_pengaduan" : kode,
                "label_asli"     : v["label_asli"],
                "prediksi_dominan": pred_dominan,
                "salah_di_fold"  : v["salah_di_fold"],
            })

    import numpy as _np
    cv_output = {
        "n_folds"           : 5,
        "rata_rata_akurasi" : round(float(_np.mean(cv_scores_list)), 4),
        "std_akurasi"       : round(float(_np.std(cv_scores_list)),  4),
        "cv_scores"         : cv_scores_list,
        "folds"             : cv_folds,
        "data_konsisten_salah": data_konsisten_salah,
    }
    save_json(cv_output, STAGE_CV_PATH)
    print(f"  [OK] cross_validation.json saved (mean={cv_output['rata_rata_akurasi']:.4f}, {len(data_konsisten_salah)} konsisten salah)")

    # Update cv_scores untuk hasil_training.json (pakai hasil fold aktual)
    import numpy as _np2
    cv_scores = _np2.array(cv_scores_list)

    # ═══════════════════════════════════════════════════════════════
    # GENERATE STAGE-SPECIFIC FILES (NORMALIZED STRUCTURE)
    # ═══════════════════════════════════════════════════════════════
    
    print("\n[*] Generating stage-specific files...")
    
    # Get all kode_pengaduan from df
    kode_list = df["kode_pengaduan"].tolist()
    
    # ── Stage 1: TF-IDF (per document) ────────────────────────────
    print("  [*] Generating tfidf.json...")
    tfidf_dict = {}
    feature_names = vectorizer.get_feature_names_out()
    
    for i, kode in enumerate(kode_list):
        tfidf_row = X_tfidf[i]
        _, term_indices = tfidf_row.nonzero()
        
        term_scores = {}
        for ti in term_indices:
            term = feature_names[ti]
            score = round(float(tfidf_row[0, ti]), 6)
            term_scores[term] = score
        
        # Sort by score descending
        tfidf_dict[kode] = dict(sorted(term_scores.items(), key=lambda x: -x[1]))
    
    save_json(tfidf_dict, STAGE_TFIDF_PATH)
    print(f"  [OK] tfidf.json saved ({len(tfidf_dict)} entries)")
    
    # ── Stage 2: Filtering (DF threshold) ─────────────────────────
    print("  [*] Generating filtering.json...")

    # Hitung DF semua term (sebelum filter) menggunakan CountVectorizer tanpa batasan
    from sklearn.feature_extraction.text import CountVectorizer as _CV
    count_vec_all = _CV(ngram_range=(1, 1))
    X_counts_all  = count_vec_all.fit_transform(df["final_text"])
    all_raw_terms = count_vec_all.get_feature_names_out()
    df_counts_all = (X_counts_all > 0).sum(axis=0).A1

    # term → df (semua term sebelum filter)
    term_df_all = {term: int(df_counts_all[i]) for i, term in enumerate(all_raw_terms)}

    # Ambil juga per-dokumen: term apa saja yang muncul di tiap dokumen (dari CountVec)
    # Digunakan untuk membangun bagian sebelum/kena_filter/tersisa per pengaduan
    # X_counts_all[i] = sparse row untuk dokumen ke-i
    doc_terms_raw = {}  # kode → list term yang muncul di dok ini (sebelum filter)
    for i, kode in enumerate(kode_list):
        _, col_idx = X_counts_all[i].nonzero()
        doc_terms_raw[kode] = [all_raw_terms[ci] for ci in col_idx]

    # Parameter filter (sama dengan TfidfVectorizer)
    min_df_threshold = 2
    max_df_threshold = int(0.95 * len(df))   # ~1140 untuk 1200 dok

    # ── Global: kumpulkan term_terbuang dan term_lolos (seluruh vocabulary) ──
    global_terbuang = []  # [{"term": ..., "df": ..., "alasan": ...}]
    global_lolos    = []  # [{"term": ..., "df": ...}]

    for term, df_val in term_df_all.items():
        if df_val < min_df_threshold:
            global_terbuang.append({"term": term, "df": df_val, "alasan": "DF < Min DF"})
        elif df_val > max_df_threshold:
            global_terbuang.append({"term": term, "df": df_val, "alasan": "DF > Max DF"})
        else:
            global_lolos.append({"term": term, "df": df_val})

    # Urutkan: terbuang by df asc (terlalu jarang dulu), lolos by df asc
    global_terbuang.sort(key=lambda x: x["df"])
    global_lolos.sort(key=lambda x: x["df"])

    # ── Per-pengaduan: sebelum/kena_filter/tersisa ─────────────────
    lolos_term_set    = {t["term"] for t in global_lolos}
    terbuang_term_set = {t["term"] for t in global_terbuang}

    filtering_dict = {}
    for kode in kode_list:
        sebelum_terms = doc_terms_raw.get(kode, [])

        kena_filter = []
        tersisa     = []
        for term in sebelum_terms:
            df_val = term_df_all.get(term, 0)
            if df_val < min_df_threshold:
                kena_filter.append({"term": term, "df": df_val, "alasan": "DF < Min DF"})
            elif df_val > max_df_threshold:
                kena_filter.append({"term": term, "df": df_val, "alasan": "DF > Max DF"})
            else:
                tersisa.append(term)

        filtering_dict[kode] = {
            "sebelum"    : sebelum_terms,
            "kena_filter": kena_filter,
            "tersisa"    : tersisa,
            "lolos_count"    : len(tersisa),
            "terbuang_count" : len(kena_filter),
        }

    # ── Gabungkan: simpan global summary + per-pengaduan dalam satu file ──
    filtering_output = {
        "_konfigurasi": {
            "min_df"       : min_df_threshold,
            "max_df_ratio" : 0.95,
            "max_df_abs"   : max_df_threshold,
            "total_dok"    : len(df),
        },
        "_global": {
            "total_sebelum_filter" : len(term_df_all),
            "total_terbuang"       : len(global_terbuang),
            "total_lolos"          : len(global_lolos),
            "term_terbuang"        : global_terbuang,
            "term_lolos"           : global_lolos,
        },
        **filtering_dict,
    }

    save_json(filtering_output, STAGE_FILTERING_PATH)
    print(f"  [OK] filtering.json saved: {len(global_lolos)} lolos, {len(global_terbuang)} terbuang, {len(filtering_dict)} entries per-pengaduan")
    
    # ── Stage 3: Seleksi Fitur (Chi-Square) ───────────────────────
    print("  [*] Generating seleksi_fitur.json...")
    seleksi_fitur_dict = {}
    
    selected_mask = selector.get_support()
    chi2_scores = selector.scores_
    
    # Map feature names to chi2 scores
    feature_chi2_dict = {}
    for i, term in enumerate(feature_names):
        feature_chi2_dict[term] = {
            "chi2_score": round(float(chi2_scores[i]), 4),
            "selected": bool(selected_mask[i])
        }
    
    for i, kode in enumerate(kode_list):
        term_terpilih = {}
        term_tidak_terpilih = {}
        
        # Get terms for this document
        tfidf_row = X_tfidf[i]
        _, term_indices = tfidf_row.nonzero()
        
        for ti in term_indices:
            term = feature_names[ti]
            chi2_score = round(float(chi2_scores[ti]), 4)
            tfidf_score = round(float(tfidf_row[0, ti]), 6)
            
            if selected_mask[ti]:
                term_terpilih[term] = {
                    "chi2_score": chi2_score,
                    "tfidf": tfidf_score
                }
            else:
                term_tidak_terpilih[term] = {
                    "chi2_score": chi2_score,
                    "tfidf": tfidf_score
                }
        
        seleksi_fitur_dict[kode] = {
            "metode": "chi-square",
            "percentile": 80,
            "term_terpilih": term_terpilih,
            "term_tidak_terpilih": term_tidak_terpilih,
            "terpilih_count": len(term_terpilih),
            "tidak_terpilih_count": len(term_tidak_terpilih)
        }
    
    save_json(seleksi_fitur_dict, STAGE_SELEKSI_FITUR_PATH)
    print(f"  [OK] seleksi_fitur.json saved ({len(seleksi_fitur_dict)} entries)")
    
    # ── Stage 3b: Bootstrap Sampling (20 pohon pertama, data nyata) ──
    print("  [*] Generating bootstrap.json (20 pohon pertama, data aktual)...")
    bootstrap_dict = {}
    
    # Reset index df agar posisi 0..N-1 konsisten dengan y_encoded
    df_reset = df.reset_index(drop=True)
    
    # Kode → nama, deskripsi, kategori (lookup untuk enrichment)
    kode_to_info = {}
    for pos, row in df_reset.iterrows():
        kode_to_info[row["kode_pengaduan"]] = {
            "nama": row.get("nama", ""),
            "deskripsi": row.get("deskripsi", ""),
            "kategori": label_encoder.inverse_transform([y_encoded[pos]])[0],
        }

    # Rekonstruksi train_idx (posisi di df_reset) dengan parameter identik saat split
    all_indices = list(range(len(df_reset)))
    from sklearn.model_selection import train_test_split as _tts
    train_idx, _ = _tts(
        all_indices, test_size=0.2, random_state=42,
        stratify=y_encoded, shuffle=True
    )
    train_kode_list  = [df_reset.iloc[i]["kode_pengaduan"] for i in train_idx]
    train_label_list = [label_encoder.inverse_transform([y_encoded[i]])[0] for i in train_idx]
    
    n_train = len(train_idx)  # biasanya 960
    
    # Ekstrak bootstrap sample tiap pohon menggunakan internal sklearn indices
    # sklearn menyimpan sample indices di tree.estimators_samples_ (hanya ada jika warm_start=False dan bootstrap=True)
    # Akses lewat model.estimators_samples_ — array of shape (n_samples,) per tree
    for tree_num in range(min(20, len(model.estimators_))):
        # estimators_samples_[i] = index array ke X_train yang dipakai pohon ke-i
        sample_indices = model.estimators_samples_[tree_num]  # shape (n_train,), bisa duplikat
        
        from collections import Counter as _Counter
        idx_counter = _Counter(sample_indices.tolist())
        
        # Data yang masuk bootstrap sample
        sampel = []
        in_bag_idx_set = set(idx_counter.keys())
        
        for orig_idx, count in sorted(idx_counter.items()):
            if orig_idx < len(train_kode_list):
                kode = train_kode_list[orig_idx]
                sampel.append({
                    "kode_pengaduan": kode,
                    "diambil": int(count),
                    "kategori": train_label_list[orig_idx],
                })
        
        # Data yang tidak masuk → OOB
        oob_kode_list = []
        for orig_idx in range(n_train):
            if orig_idx not in in_bag_idx_set and orig_idx < len(train_kode_list):
                oob_kode_list.append(train_kode_list[orig_idx])
        
        # Distribusi kelas di bootstrap sample ini
        class_dist = {}
        for s in sampel:
            cat = s["kategori"]
            class_dist[cat] = class_dist.get(cat, 0) + s["diambil"]
        
        # Distribusi kelas OOB
        oob_class_dist = {}
        for orig_idx in range(n_train):
            if orig_idx not in in_bag_idx_set and orig_idx < len(train_label_list):
                cat = train_label_list[orig_idx]
                oob_class_dist[cat] = oob_class_dist.get(cat, 0) + 1
        
        bootstrap_dict[f"tree_{tree_num + 1}"] = {
            "total_sampel": len(sample_indices),
            "unique_sampel": len(in_bag_idx_set),
            "duplikat": len(sample_indices) - len(in_bag_idx_set),
            "oob_count": len(oob_kode_list),
            "class_distribution": class_dist,
            "oob_class_distribution": oob_class_dist,
            "sampel": sampel,
            "oob": oob_kode_list,
        }
    
    # Simpan kode_to_info sebagai lookup terpisah (lebih efisien daripada duplikasi di tiap pohon)
    save_json(kode_to_info, os.path.join(STAGES_DIR, "bootstrap_lookup.json"))
    save_json(bootstrap_dict, os.path.join(STAGES_DIR, "bootstrap.json"))
    print(f"  [OK] bootstrap.json saved ({len(bootstrap_dict)} pohon)")
    print(f"  [OK] bootstrap_lookup.json saved ({len(kode_to_info)} entri)")

    # ── Stage 3c: Gini Impurity & Node Splitting (20 pohon pertama) ──
    print("  [*] Generating gini_splitting.json (20 pohon, node aktual)...")
    os.makedirs(STAGE_GINI_SAMPLES_DIR, exist_ok=True)

    # Mapping: index fitur setelah seleksi → nama term
    selected_feature_names = [feature_names[i] for i in range(len(feature_names)) if selected_mask[i]]

    # Rekonstruksi train_idx (sudah ada dari bootstrap generation)
    # _train_idx_oob & _train_kode (sudah didefinisikan di bagian OOB di atas)
    # Kita pakai X_train dan train_idx dari split utama
    from sklearn.model_selection import train_test_split as _tts3
    _all_idx3 = list(range(len(df.reset_index(drop=True))))
    _train_idx3, _ = _tts3(
        _all_idx3, test_size=0.2, random_state=42,
        stratify=y_encoded, shuffle=True
    )
    _df3         = df.reset_index(drop=True)
    _train_kode3 = [_df3.iloc[i]["kode_pengaduan"] for i in _train_idx3]

    CATS_ORDER = label_encoder.classes_.tolist()  # ['INFRASTRUKTUR', 'KEAMANAN', 'LINGKUNGAN', 'PELAYANAN'] (sorted)
    N_GINI_TREES = min(20, len(model.estimators_))

    gini_output    = {}
    term_split_agg = {}  # term → {frekuensi, pohon_ids, gini_values}

    for tree_num in range(min(N_GINI_TREES, len(model.estimators_))):
        tree        = model.estimators_[tree_num]
        sk_tree     = tree.tree_
        n_nodes     = sk_tree.node_count
        n_train_samples = len(_train_idx3)

        # Dapatkan sample indices yang dipakai pohon ini (bootstrap)
        samp_idx_boot = model.estimators_samples_[tree_num]

        # decision_path untuk semua training samples → tahu tiap sampel ada di node mana
        # Gunakan X_train (full train set) bukan X_selected agar sesuai bootstrap
        try:
            node_indicator = tree.decision_path(X_train)  # sparse matrix (n_train, n_nodes)
        except Exception:
            node_indicator = None

        # Build: node_id → list of train-sample indices yang jatuh ke node ini
        node_to_samples = {}  # node_id (int) → list of index ke _train_kode3
        if node_indicator is not None:
            node_csr = node_indicator.tocsr()
            for sample_rank in range(node_csr.shape[0]):
                for node_id in node_csr.getrow(sample_rank).indices:
                    if node_id not in node_to_samples:
                        node_to_samples[node_id] = []
                    node_to_samples[node_id].append(sample_rank)

        nodes_list = []
        for node_id in range(n_nodes):
            is_leaf    = (sk_tree.children_left[node_id] == -1)
            depth      = int(sk_tree.compute_node_depths()[node_id]) if hasattr(sk_tree, "compute_node_depths") else -1
            gini_val   = round(float(sk_tree.impurity[node_id]), 6)
            n_samp     = int(sk_tree.n_node_samples[node_id])

            # Distribusi kelas di node ini (dari sk_tree.value)
            # sk_tree.value shape: (n_nodes, n_outputs, n_classes)
            # Dengan class_weight, value berisi proporsi — kalikan n_node_samples untuk count nyata
            class_values_raw = sk_tree.value[node_id][0]
            total_val = class_values_raw.sum()
            if total_val > 0 and abs(total_val - 1.0) < 0.01:
                # Proporsi — konversi ke count
                class_counts = np.round(class_values_raw * n_samp).astype(int)
            else:
                class_counts = class_values_raw.astype(int)
            distribusi   = {CATS_ORDER[ci]: int(class_counts[ci]) for ci in range(len(CATS_ORDER))}
            prediksi_node = CATS_ORDER[int(class_counts.argmax())]

            # Term split
            term_split  = None
            threshold   = None
            child_kiri  = None
            child_kanan = None

            if not is_leaf:
                feat_idx = int(sk_tree.feature[node_id])
                if 0 <= feat_idx < len(selected_feature_names):
                    term_split = selected_feature_names[feat_idx]
                threshold   = round(float(sk_tree.threshold[node_id]), 6)
                child_kiri  = int(sk_tree.children_left[node_id])
                child_kanan = int(sk_tree.children_right[node_id])

                # Akumulasi term split global
                if term_split:
                    if term_split not in term_split_agg:
                        term_split_agg[term_split] = {"frekuensi": 0, "pohon_ids": [], "gini_values": []}
                    term_split_agg[term_split]["frekuensi"] += 1
                    if (tree_num + 1) not in term_split_agg[term_split]["pohon_ids"]:
                        term_split_agg[term_split]["pohon_ids"].append(tree_num + 1)
                    term_split_agg[term_split]["gini_values"].append(gini_val)

            # Kode pengaduan yang ada di node ini (for on-demand lookup)
            # Simpan sebagai file terpisah per (tree, node) agar gini_splitting.json tidak terlalu besar
            sampel_kode = []
            if node_id in node_to_samples:
                for sr in node_to_samples[node_id]:
                    if sr < len(_train_kode3):
                        sampel_kode.append(_train_kode3[sr])

            node_entry = {
                "node_id"     : node_id,
                "kedalaman"   : depth,
                "tipe"        : "leaf" if is_leaf else "split",
                "term_split"  : term_split,
                "threshold"   : threshold,
                "gini"        : gini_val,
                "is_pure"     : gini_val <= 0.05,
                "jumlah_sampel": n_samp,
                "distribusi"  : distribusi,
                "prediksi"    : prediksi_node,
                "child_kiri"  : child_kiri,
                "child_kanan" : child_kanan,
                "n_sampel_ids": len(sampel_kode),  # jumlah saja, data aslinya di file terpisah
            }
            nodes_list.append(node_entry)

            # Simpan sampel_ids ke file terpisah (on-demand)
            if sampel_kode:
                sampel_file = os.path.join(STAGE_GINI_SAMPLES_DIR, f"tree{tree_num+1}_node{node_id}.json")
                save_json(sampel_kode, sampel_file)

        # Hitung statistik pohon
        split_nodes = [n for n in nodes_list if n["tipe"] == "split"]
        leaf_nodes  = [n for n in nodes_list if n["tipe"] == "leaf"]
        max_depth   = max((n["kedalaman"] for n in nodes_list if n["kedalaman"] >= 0), default=0)
        gini_values = [n["gini"] for n in nodes_list if n["tipe"] == "split"]
        rata_gini   = round(float(sum(gini_values) / len(gini_values)), 4) if gini_values else 0.0

        # Top term split di pohon ini (term paling atas/dekat root)
        top_term = None
        if nodes_list and nodes_list[0]["term_split"]:
            top_term = nodes_list[0]["term_split"]

        gini_output[f"tree_{tree_num + 1}"] = {
            "total_node"  : n_nodes,
            "total_split" : len(split_nodes),
            "total_leaf"  : len(leaf_nodes),
            "kedalaman"   : max_depth,
            "rata_gini"   : rata_gini,
            "top_term"    : top_term,
            "nodes"       : nodes_list,
        }

    # Finalisasi term_split_global: hitung rata_gini dan urutkan by frekuensi
    term_split_global = []
    for term, v in term_split_agg.items():
        avg_g = round(float(sum(v["gini_values"]) / len(v["gini_values"])), 4) if v["gini_values"] else 0.0
        term_split_global.append({
            "term"       : term,
            "frekuensi"  : v["frekuensi"],
            "pohon_unik" : len(v["pohon_ids"]),
            "pohon_ids"  : sorted(v["pohon_ids"])[:20],  # max 20 pohon ids
            "rata_gini"  : avg_g,
        })
    term_split_global.sort(key=lambda x: -x["pohon_unik"])

    gini_output["_meta"] = {
        "n_trees"         : N_GINI_TREES,
        "n_features"      : len(selected_feature_names),
        "n_train"         : len(_train_idx3),
        "samples_dir"     : "stages/gini_samples",
    }
    gini_output["_term_split_global"] = term_split_global

    save_json(gini_output, STAGE_GINI_PATH)
    print(f"  [OK] gini_splitting.json saved ({N_GINI_TREES} pohon, {len(term_split_global)} unique split terms)")
    print("  [*] Generating random_forest.json...")
    random_forest_dict = {}
    
    # Get feature importances
    feature_importances = model.feature_importances_.tolist()
    
    # Map selected features to importances
    selected_features = [feature_names[i] for i in range(len(feature_names)) if selected_mask[i]]
    feature_importance_dict = {selected_features[i]: round(float(feature_importances[i]), 6) 
                               for i in range(min(len(selected_features), len(feature_importances)))}
    
    # Get predictions for all data
    all_pred  = model.predict(X_selected)
    all_proba = model.predict_proba(X_selected)
    
    # Per-tree predictions untuk 20 pohon pertama (cukup untuk voting transparency)
    N_VOTE_TREES = min(20, len(model.estimators_))
    tree_preds_20 = []  # shape (N_VOTE_TREES, n_samples)
    for t_idx in range(min(N_VOTE_TREES, len(model.estimators_))):
        tp = model.estimators_[t_idx].predict(X_selected)
        tree_preds_20.append(tp)

    # OOB membership: kode → set of tree_idx yang TIDAK memasukkan sampel ini ke bootstrap
    # Gunakan estimators_samples_ untuk X_train (lebih akurat)
    # Kita perlu mapping kode → posisi di X_selected (kode_list[i] = kode)
    kode_to_xsel_idx = {kode: i for i, kode in enumerate(kode_list)}

    # Rekonstruksi train_idx (untuk OOB check)
    from sklearn.model_selection import train_test_split as _tts_v
    _df_v   = df.reset_index(drop=True)
    _all_v  = list(range(len(_df_v)))
    _train_idx_v, _ = _tts_v(_all_v, test_size=0.2, random_state=42, stratify=y_encoded, shuffle=True)
    _train_kode_v   = [_df_v.iloc[i]["kode_pengaduan"] for i in _train_idx_v]
    _train_kode_set = set(_train_kode_v)

    # Untuk setiap pohon dalam 20 pohon, simpan indeks yang masuk bootstrap (in-bag)
    inbag_per_tree = []
    for t_idx in range(min(N_VOTE_TREES, len(model.estimators_))):
        samp_idx = model.estimators_samples_[t_idx]
        inbag_set = set(samp_idx.tolist())
        inbag_per_tree.append(inbag_set)

    for i, kode in enumerate(kode_list):
        rf_data = {
            "prediction": label_encoder.inverse_transform([int(all_pred[i])])[0],
            "confidence": round(float(all_proba[i].max()), 4),
            "proba_all": {cat: round(float(all_proba[i, j]), 4) 
                          for j, cat in enumerate(label_encoder.classes_)},
        }
        
        # Per-tree votes (20 pohon)
        tree_votes_dict = {}
        for t_idx, tp in enumerate(tree_preds_20):
            vote = label_encoder.inverse_transform([int(tp[i])])[0]
            # Cek apakah kode ini OOB di pohon ini
            # Kode harus ada di train set, dan posisi train-nya tidak ada di inbag pohon ini
            is_train = kode in _train_kode_set
            is_oob   = False
            if is_train:
                # Cari posisi di train_idx
                try:
                    pos_in_train = _train_kode_v.index(kode)
                    is_oob = pos_in_train not in inbag_per_tree[t_idx]
                except ValueError:
                    pass
            tree_votes_dict[f"tree_{t_idx + 1}"] = {
                "vote"   : vote,
                "is_oob" : bool(is_oob),
            }
        rf_data["tree_votes_sample"] = tree_votes_dict
        rf_data["total_trees"] = len(model.estimators_)
        rf_data["oob"] = False  # legacy placeholder
        
        # Feature importance contribution (top features for this document)
        doc_features = {}
        tfidf_row = X_tfidf[i]
        _, term_indices = tfidf_row.nonzero()
        
        for ti in term_indices:
            if selected_mask[ti]:
                term = feature_names[ti]
                if term in feature_importance_dict:
                    doc_features[term] = {
                        "importance": feature_importance_dict[term],
                        "tfidf": round(float(tfidf_row[0, ti]), 6)
                    }
        
        doc_features = dict(sorted(doc_features.items(), 
                                   key=lambda x: x[1]["importance"] * x[1]["tfidf"], 
                                   reverse=True)[:20])
        rf_data["feature_importance_kontribusi"] = doc_features
        
        random_forest_dict[kode] = rf_data
    
    save_json(random_forest_dict, STAGE_RANDOM_FOREST_PATH)
    print(f"  [OK] random_forest.json saved ({len(random_forest_dict)} entries)")

    # ── Majority Voting transparency file ──────────────────────────
    print("  [*] Generating majority_voting.json...")

    # Mapping kode → label asli (dari df)
    label_col_v = "Kategori" if "Kategori" in _df_v.columns else "kategori"
    kode_to_label = {}
    for idx_r, row_r in _df_v.iterrows():
        k_r = row_r.get("kode_pengaduan")
        l_r = label_encoder.inverse_transform([y_encoded[idx_r]])[0]
        kode_to_label[k_r] = l_r

    # selected_feature_names sudah ada dari bagian gini di atas, tapi kalau tidak ada:
    if "selected_feature_names" not in dir():
        selected_feature_names = [feature_names[j] for j in range(len(feature_names)) if selected_mask[j]]

    from collections import Counter as _C_v

    voting_dict = {}
    for i, kode in enumerate(kode_list):
        rf_entry = random_forest_dict[kode]
        tv       = rf_entry.get("tree_votes_sample", {})

        # Distribusi vote dari 20 pohon
        all_votes     = [v["vote"] for v in tv.values()]
        dist_vote     = dict(_C_v(all_votes))
        majority_vote = rf_entry["prediction"]
        confidence    = rf_entry["confidence"]

        # Label asli
        label_asli = kode_to_label.get(kode, "-")
        benar      = (majority_vote == label_asli)

        # Pohon representatif: 2 pohon yang vote ke majority + 1 yang tidak (jika ada)
        majority_trees   = [k for k, v in tv.items() if v["vote"] == majority_vote]
        minority_trees   = [k for k, v in tv.items() if v["vote"] != majority_vote]
        rep_trees_keys   = majority_trees[:2] + minority_trees[:1]

        pohon_representatif = {}
        for tree_key in rep_trees_keys:
            t_num  = int(tree_key.replace("tree_", "")) - 1
            if t_num >= len(model.estimators_):
                continue
            tree_sk  = model.estimators_[t_num]
            sk_tree  = tree_sk.tree_

            # Ikuti jalur keputusan untuk sampel ini
            x_row = X_selected[i]
            try:
                node_path = tree_sk.decision_path(x_row).indices.tolist()
            except Exception:
                node_path = []

            jalur = []
            for node_id in node_path:
                is_leaf = (sk_tree.children_left[node_id] == -1)
                feat_idx = int(sk_tree.feature[node_id]) if not is_leaf else -1
                thresh   = round(float(sk_tree.threshold[node_id]), 6) if not is_leaf else None
                gini_n   = round(float(sk_tree.impurity[node_id]), 4)

                # Tentukan arah (kiri atau kanan) dari node ini
                arah = None
                if not is_leaf and len(node_path) > node_path.index(node_id) + 1:
                    next_node = node_path[node_path.index(node_id) + 1]
                    arah = "kiri" if sk_tree.children_left[node_id] == next_node else "kanan"

                term = selected_feature_names[feat_idx] if (feat_idx >= 0 and feat_idx < len(selected_feature_names)) else None

                if is_leaf:
                    class_vals = sk_tree.value[node_id][0]
                    pred_node  = label_encoder.inverse_transform([int(class_vals.argmax())])[0]
                    jalur.append({
                        "node_id" : node_id,
                        "tipe"    : "leaf",
                        "prediksi": pred_node,
                        "gini"    : gini_n,
                    })
                else:
                    jalur.append({
                        "node_id"  : node_id,
                        "tipe"     : "split",
                        "term"     : term,
                        "threshold": thresh,
                        "arah"     : arah,
                        "gini"     : gini_n,
                    })

            pohon_representatif[tree_key] = {
                "vote"        : tv[tree_key]["vote"],
                "is_oob"      : tv[tree_key]["is_oob"],
                "jalur_node"  : jalur,
            }

        voting_dict[kode] = {
            "vote_per_pohon"      : tv,
            "distribusi_vote"     : dist_vote,
            "majority_vote"       : majority_vote,
            "confidence"          : confidence,
            "label_asli"          : label_asli,
            "benar"               : benar,
            "pohon_representatif" : pohon_representatif,
        }

    save_json(voting_dict, STAGE_VOTING_PATH)
    print(f"  [OK] majority_voting.json saved ({len(voting_dict)} entries)")
    
    # ═══════════════════════════════════════════════════════════════
    # GENERATE FINAL_PROCESSED.JSON (SLIM VERSION)
    # ═══════════════════════════════════════════════════════════════
    
    print("\n[*] Generating final_processed.json (slim version)...")
    hasil_semua = []
    data_list = df.to_dict("records")
    
    for i, row in enumerate(data_list):
        hasil_semua.append({
            "kode_pengaduan"    : row.get("kode_pengaduan"),
            "nama"              : row.get("nama", ""),
            "no_wa"             : str(row.get("no_wa", "")).replace("@c.us", ""),
            "deskripsi"         : row.get("deskripsi", ""),
            "processed"         : row.get("final_text", ""),
            "label_asli"        : label_encoder.inverse_transform([int(y_encoded[i])])[0],
            "kategori_prediksi" : label_encoder.inverse_transform([int(all_pred[i])])[0],
            "confidence"        : round(float(all_proba[i].max()), 4),
            "timestamp"         : row.get("timestamp", "-"),
        })
    
    save_json(hasil_semua, FINAL_PROCESSED_PATH)
    print(f"[OK] final_processed.json saved ({len(hasil_semua)} entries)")

    # ── Export TF-IDF term data untuk halaman TF-IDF ──────────────
    print("[*] Menyimpan data TF-IDF terms...")
    selected_mask   = selector.get_support()                    # bool array 2612 -> 1000 True
    chi2_scores     = selector.scores_                          # chi2 score tiap term

    # Hitung rata-rata TF-IDF per term dari X_tfidf (full matrix, semua 1200 dok)
    tfidf_mean      = X_tfidf.mean(axis=0).A1                  # shape (2612,)

    tfidf_terms = []
    for idx, term in enumerate(feature_names):
        tfidf_terms.append({
            "term"       : term,
            "ngram"      : "unigram",
            "tfidf_mean" : round(float(tfidf_mean[idx]), 6),
            "chi2_score" : round(float(chi2_scores[idx]), 4),
            "selected"   : bool(selected_mask[idx]),
        })

    # Urutkan: terpilih dulu, lalu chi2 score turun
    tfidf_terms.sort(key=lambda x: (-int(x["selected"]), -x["chi2_score"]))

    # Simpan top 3000 (agar file tidak terlalu besar)
    save_json(tfidf_terms[:3000], os.path.join(DATA_DIR, "tfidf_terms.json"))
    print(f"[OK] {len(tfidf_terms[:3000])} TF-IDF terms disimpan ke: tfidf_terms.json")

    # ── Simpan beberapa sample dokumen ter-vektorisasi ────────────
    print("[*] Menyimpan sample dokumen TF-IDF...")
    sample_indices  = [0, 1, 2, 3, 4]            # ambil 5 dokumen pertama
    sample_fn_names = feature_names               # semua term sebelum seleksi
    doc_samples     = []

    for si in sample_indices:
        if si >= len(hasil_semua):
            break
        row       = hasil_semua[si]
        tfidf_row = X_tfidf[si]                  # sparse row untuk dokumen ini

        # Ambil term yang nilainya > 0 untuk dokumen ini
        _, term_indices  = tfidf_row.nonzero()
        term_scores      = [(sample_fn_names[ti], round(float(tfidf_row[0, ti]), 6))
                            for ti in term_indices]
        term_scores.sort(key=lambda x: -x[1])    # urut bobot terbesar dulu

        doc_samples.append({
            "nama"       : row.get("nama", ""),
            "deskripsi"  : row.get("deskripsi", ""),
            "processed"  : row.get("processed", ""),
            "label_asli" : row.get("label_asli", "-"),
            "kategori"   : row.get("kategori_prediksi", "-"),
            "terms"      : term_scores[:30],     # top-30 term per dokumen
        })

    save_json(doc_samples, os.path.join(DATA_DIR, "tfidf_sample_docs.json"))
    print(f"[OK] {len(doc_samples)} sample dokumen disimpan ke: tfidf_sample_docs.json")

    # ── Generate tfidf_matrix.json (matriks input Random Forest) ──
    print("[*] Generating tfidf_matrix.json (matriks X_selected per pengaduan)...")

    # X_selected: matriks (n_dok × n_term) — nilai TF-IDF setelah SelectPercentile
    # Ini adalah matriks yang BENAR-BENAR masuk ke RandomForest
    # selected_feature_names sudah tersedia dari bagian sebelumnya
    if "selected_feature_names" not in dir():
        selected_feature_names = [feature_names[j] for j in range(len(feature_names)) if selected_mask[j]]

    # Meta: daftar term + chi2 score
    chi2_scores_arr = selector.scores_
    selected_indices = [j for j in range(len(feature_names)) if selected_mask[j]]
    terms_meta = []
    for rank, j in enumerate(selected_indices):
        terms_meta.append({
            "term"       : feature_names[j],
            "chi2_score" : round(float(chi2_scores_arr[j]), 4),
            "ngram"      : "unigram",
        })
    # Urutkan by chi2 descending untuk tampilan
    terms_meta.sort(key=lambda x: -x["chi2_score"])
    selected_term_order = [t["term"] for t in terms_meta]  # urutan term untuk matriks

    # Buat mapping term → posisi di X_selected (X_selected sudah diurutkan oleh selector)
    # Kolom X_selected[i, c] = TF-IDF term ke-c dalam selected_feature_names
    term_to_col = {term: c for c, term in enumerate(selected_feature_names)}

    # Bangun matriks per kode_pengaduan (hanya simpan nilai > 0 untuk efisiensi)
    tfidf_matrix = {}
    for i, kode in enumerate(kode_list):
        row = X_selected[i]
        # Dapatkan nilai per term (sparse row)
        _, col_idx = row.nonzero()
        term_vals = {}
        for c in col_idx:
            term_name = selected_feature_names[c]
            term_vals[term_name] = round(float(row[0, c]), 6)
        tfidf_matrix[kode] = term_vals

    # Simpan dalam format yang dioptimalkan untuk lazy loading
    tfidf_matrix_output = {
        "meta": {
            "total_pengaduan" : len(kode_list),
            "total_term"      : len(selected_feature_names),
            "metode_seleksi"  : "SelectPercentile chi-squared",
            "terms"           : selected_term_order,  # urutan by chi2 desc
            "terms_chi2"      : terms_meta,
        },
        "matrix": tfidf_matrix,
    }
    tfidf_matrix_path = os.path.join(DATA_DIR, "tfidf_matrix.json")
    save_json(tfidf_matrix_output, tfidf_matrix_path)
    print(f"[OK] tfidf_matrix.json disimpan ({len(kode_list)} pengaduan × {len(selected_feature_names)} term)")
    export_preprocessing_to_excel(load_json(DATASET_BERLABEL_PATH), hasil_semua)

# =====================================================
# 7. PREDIKSI DATA BARU
# =====================================================
def _make_dedup_key(item):
    """Buat kunci unik dari kombinasi no_wa + deskripsi + timestamp."""
    return (
        str(item.get("no_wa", "")).strip(),
        str(item.get("deskripsi", "")).strip(),
        str(item.get("timestamp", "")).strip(),
    )


def predict_new_data():
    print("[ ] Memproses data baru...")

    if not os.path.exists(DATA_BARU_PATH):
        save_json([], HASIL_PREDIKSI_PATH)
        print("[!] data_baru.json tidak ditemukan")
        return

    data_baru = load_json(DATA_BARU_PATH)

    if len(data_baru) == 0:
        save_json([], HASIL_PREDIKSI_PATH)
        print("[*] data_baru.json kosong")
        return

    print(f"[*] Jumlah data baru ditemukan: {len(data_baru)}")

    # ── 1. Preprocessing + append ke history ──────────────────
    print(f"[ ] Menambahkan {len(data_baru)} data ke preprocessing history...")
    data_preprocessed = preprocess_pipeline(data_baru, save_files=True, append_files=True)
    print(f"[OK] Preprocessing history diperbarui.")

    # ── 2. Klasifikasi ─────────────────────────────────────────
    print("[ ] Melakukan klasifikasi...")
    model         = joblib.load(MODEL_PATH)
    vectorizer    = joblib.load(VECTORIZER_PATH)
    selector      = joblib.load(SELECTOR_PATH)
    label_encoder = joblib.load(LABEL_ENCODER_PATH)

    df                = pd.DataFrame(data_preprocessed)
    X_text            = df["final_text"]
    X_tfidf           = vectorizer.transform(X_text)
    X_selected        = selector.transform(X_tfidf)
    predictions       = model.predict(X_selected)
    predicted_labels  = label_encoder.inverse_transform(predictions)
    probabilities     = model.predict_proba(X_selected)
    confidence_scores = probabilities.max(axis=1)
    categories        = label_encoder.classes_.tolist()

    # Bangun list hasil prediksi baru (format lengkap)
    hasil_baru = []
    for i, item in enumerate(data_baru):
        # Build full probability dict for all categories
        proba_all = {}
        for j, cat in enumerate(categories):
            proba_all[cat] = round(float(probabilities[i, j]), 4)
        
        hasil_baru.append({
            "nama"              : item.get("nama", ""),
            "no_wa"             : str(item.get("no_wa", "")).replace("@c.us", ""),
            "deskripsi"         : item.get("deskripsi", ""),
            "processed"         : data_preprocessed[i].get("final_text", ""),
            "label_asli"        : "-",
            "kategori_prediksi" : predicted_labels[i],
            "confidence"        : round(float(confidence_scores[i]), 4),
            "proba_all"         : proba_all,
            "timestamp"         : item.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        })

    # Simpan ke hasil_prediksi.json (sesi ini saja, tidak di-append)
    save_json(hasil_baru, HASIL_PREDIKSI_PATH)
    print(f"[OK] {len(hasil_baru)} hasil prediksi disimpan ke: {HASIL_PREDIKSI_PATH}")

    # ── 3. Append ke final_processed.json (dengan dedup) ──────
    print(f"[ ] Menambahkan {len(hasil_baru)} data ke final_processed.json...")

    existing_final = []
    if os.path.exists(FINAL_PROCESSED_PATH):
        existing_final = load_json(FINAL_PROCESSED_PATH)

    # Bangun set kunci dari data yang sudah ada
    existing_keys = {_make_dedup_key(rec) for rec in existing_final}

    ditambahkan = 0
    diskip      = 0
    for rec in hasil_baru:
        key = _make_dedup_key(rec)
        if key in existing_keys:
            diskip += 1
        else:
            existing_final.append(rec)
            existing_keys.add(key)
            ditambahkan += 1

    save_json(existing_final, FINAL_PROCESSED_PATH)

    if diskip > 0:
        print(f"[*] {diskip} data duplikat dilewati.")
    print(f"[ ] Menambahkan {ditambahkan} data ke final_processed.json")

    total_pengaduan = len(existing_final)
    print(f"[*] Total pengaduan saat ini: {total_pengaduan}")
    print("[OK] Sinkronisasi selesai")

    # ── 4. Kosongkan data_baru.json agar tidak diklasifikasi ulang ──
    save_json([], DATA_BARU_PATH)
    print("[OK] data_baru.json dikosongkan.")

# =====================================================
# 8. MAIN
# =====================================================
if __name__ == "__main__":
    print(f"[*] Project Root: {PROJECT_ROOT}")

    if not os.path.exists(DATASET_BERLABEL_PATH):
        print(f"[ERROR] Dataset berlabel tidak ditemukan: {DATASET_BERLABEL_PATH}")
        sys.exit(1)

    force_train   = "--train"  in sys.argv
    export_only   = "--export" in sys.argv
    classify_only = "--classify" in sys.argv

    if export_only:
        print("[*] Mode export Excel...")
        if os.path.exists(FINAL_PROCESSED_PATH):
            final_data = load_json(FINAL_PROCESSED_PATH)
            raw_data   = load_json(DATASET_BERLABEL_PATH)
            export_preprocessing_to_excel(raw_data, final_data)
        else:
            print("[!] final_processed.json belum ada. Jalankan training terlebih dahulu.")
            sys.exit(1)
    elif classify_only:
        print("[*] Mode klasifikasi real-time...")
        if not os.path.exists(MODEL_PATH):
            print(f"[ERROR] Model belum tersedia: {MODEL_PATH}")
            print("[INFO] Jalankan: python main.py --train")
            sys.exit(1)
        predict_new_data()
    elif not os.path.exists(MODEL_PATH) or force_train:
        print("[*] Force retrain..." if force_train else "[*] Model belum ada. Melatih...")
        train_model()
        predict_new_data()
    else:
        print("[OK] Model sudah tersedia.")
        predict_new_data()

    print("[OK] Program selesai dijalankan.")
